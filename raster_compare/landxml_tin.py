from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping
import csv
import hashlib
import json
import math
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


DEFAULT_CONFIG: dict[str, Any] = {
    "excel": True,
    "surface_name": None,
    "coordinate_order": "xyz",
    "outputs": {
        "vertices_csv": "vertices.csv",
        "faces_csv": "faces.csv",
        "obj": "tin_mesh.obj",
        "ply": "tin_mesh.ply",
        "excel_report": "landxml_tin_report.xlsx",
        "summary_json": "landxml_tin_summary.json",
    },
    "options": {
        "write_ply": True,
        "strict_faces": True,
        "validate_counts": True,
        "preserve_original_point_ids": True,
        "degenerate_area_tolerance": 1e-12,
    },
    "expected": {
        "surface_type": "TIN",
        "n_points": None,
        "n_faces": None,
    },
}


@dataclass(frozen=True)
class TinSurface:
    name: str | None
    surface_type: str | None
    vertices: dict[str, tuple[float, float, float]]
    faces: list[tuple[str, str, str]]
    available_surfaces: list[dict[str, Any]]
    warnings: list[str]
    coordinate_order: str = "xyz"


def _deep_update(base: dict[str, Any], updates: Mapping[str, Any]) -> dict[str, Any]:
    for key, value in updates.items():
        if isinstance(value, Mapping) and isinstance(base.get(key), dict):
            _deep_update(base[key], value)  # type: ignore[index]
        else:
            base[key] = value
    return base


def _copy_defaults(data: Mapping[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in data.items():
        if isinstance(value, Mapping):
            out[key] = _copy_defaults(value)
        else:
            out[key] = value
    return out


def _flatten_mapping(data: Mapping[str, Any], prefix: str = "") -> Iterable[tuple[str, Any]]:
    for key, value in data.items():
        full_key = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, Mapping):
            yield from _flatten_mapping(value, full_key)
        else:
            yield full_key, value


def _local_name(tag: str) -> str:
    return tag.split("}", 1)[-1] if "}" in tag else tag


def _iter_children(elem: ET.Element, local_name: str) -> Iterable[ET.Element]:
    for child in list(elem):
        if _local_name(child.tag) == local_name:
            yield child


def _find_first_child(elem: ET.Element, local_name: str) -> ET.Element | None:
    for child in _iter_children(elem, local_name):
        return child
    return None


def _validate_coordinate_order(coordinate_order: str) -> str:
    order = str(coordinate_order).lower().strip()
    if sorted(order) != ["x", "y", "z"] or len(order) != 3:
        raise ValueError("coordinate_order must be a 3-character permutation of 'xyz', e.g. 'xyz' or 'yxz'.")
    return order


def _apply_coordinate_order(values: tuple[float, float, float], coordinate_order: str) -> tuple[float, float, float]:
    mapping = {axis: values[idx] for idx, axis in enumerate(coordinate_order)}
    return mapping["x"], mapping["y"], mapping["z"]


def _parse_float_triplet(text: str | None, context: str, coordinate_order: str = "xyz") -> tuple[float, float, float]:
    if text is None:
        raise ValueError(f"Missing coordinate text in {context}.")
    parts = text.strip().split()
    if len(parts) < 3:
        raise ValueError(f"Expected at least 3 coordinate values in {context}; got: {text!r}")
    try:
        raw = (float(parts[0]), float(parts[1]), float(parts[2]))
    except ValueError as exc:
        raise ValueError(f"Invalid coordinate values in {context}: {text!r}") from exc
    return _apply_coordinate_order(raw, coordinate_order)


def _parse_face_ids(text: str | None, context: str) -> tuple[str, str, str]:
    if text is None:
        raise ValueError(f"Missing face text in {context}.")
    parts = text.strip().split()
    if len(parts) != 3:
        raise ValueError(f"Only triangular TIN faces are supported in V1; {context} has {len(parts)} vertices.")
    return str(parts[0]), str(parts[1]), str(parts[2])


def _surface_type(surface: ET.Element, definition: ET.Element | None) -> str | None:
    candidates: list[str | None] = []
    if definition is not None:
        candidates.extend(
            [
                definition.attrib.get("surfType"),
                definition.attrib.get("surfType".lower()),
                definition.attrib.get("type"),
                definition.attrib.get("Type"),
            ]
        )
    candidates.extend([surface.attrib.get("surfType"), surface.attrib.get("type"), surface.attrib.get("Type")])
    for value in candidates:
        if value:
            return str(value)
    return None


def _is_tin(surface_type: str | None) -> bool:
    return bool(surface_type and str(surface_type).upper() == "TIN")


def _find_surfaces(root: ET.Element) -> list[ET.Element]:
    return [elem for elem in root.iter() if _local_name(elem.tag) == "Surface"]


def _surface_summary(surface: ET.Element) -> dict[str, Any]:
    definition = _find_first_child(surface, "Definition")
    pnts = _find_first_child(definition, "Pnts") if definition is not None else None
    faces = _find_first_child(definition, "Faces") if definition is not None else None
    return {
        "name": surface.attrib.get("name"),
        "surface_type": _surface_type(surface, definition),
        "n_points": sum(1 for _ in _iter_children(pnts, "P")) if pnts is not None else 0,
        "n_faces": sum(1 for _ in _iter_children(faces, "F")) if faces is not None else 0,
    }


def list_landxml_surfaces(xml_path: Path) -> list[dict[str, Any]]:
    root = ET.parse(xml_path).getroot()
    return [_surface_summary(surface) for surface in _find_surfaces(root)]


def _select_surface(root: ET.Element, surface_name: str | None) -> tuple[ET.Element, list[dict[str, Any]], list[str]]:
    surfaces = _find_surfaces(root)
    summaries = [_surface_summary(surface) for surface in surfaces]
    warnings: list[str] = []

    if not surfaces:
        raise ValueError("No LandXML Surface elements found.")

    if surface_name:
        for surface, summary in zip(surfaces, summaries):
            if summary.get("name") == surface_name:
                return surface, summaries, warnings
        names = ", ".join(str(s.get("name")) for s in summaries)
        raise ValueError(f"Surface named {surface_name!r} not found. Available surfaces: {names}")

    tin_surfaces = [(surface, summary) for surface, summary in zip(surfaces, summaries) if _is_tin(summary.get("surface_type"))]
    if not tin_surfaces:
        names = ", ".join(f"{s.get('name')} ({s.get('surface_type')})" for s in summaries)
        raise ValueError(f"No TIN surface found. Available surfaces: {names}")
    if len(tin_surfaces) > 1:
        warnings.append("Multiple TIN surfaces found; selected the first one. Set surface_name for explicit selection.")
    return tin_surfaces[0][0], summaries, warnings


def extract_tin_surface(xml_path: Path, surface_name: str | None = None, coordinate_order: str = "xyz") -> TinSurface:
    coordinate_order = _validate_coordinate_order(coordinate_order)
    root = ET.parse(xml_path).getroot()
    surface, summaries, warnings = _select_surface(root, surface_name)
    definition = _find_first_child(surface, "Definition")
    if definition is None:
        raise ValueError(f"Surface {surface.attrib.get('name')!r} has no Definition element.")

    surface_type = _surface_type(surface, definition)
    if not _is_tin(surface_type):
        warnings.append(f"Selected surface type is {surface_type!r}; expected TIN.")

    pnts = _find_first_child(definition, "Pnts")
    faces_elem = _find_first_child(definition, "Faces")
    if pnts is None:
        raise ValueError(f"Surface {surface.attrib.get('name')!r} has no Pnts element.")
    if faces_elem is None:
        raise ValueError(f"Surface {surface.attrib.get('name')!r} has no Faces element.")

    vertices: dict[str, tuple[float, float, float]] = {}
    duplicate_ids: list[str] = []
    for idx, p in enumerate(_iter_children(pnts, "P"), start=1):
        point_id = p.attrib.get("id") or p.attrib.get("ID") or p.attrib.get("name") or str(idx)
        point_id = str(point_id)
        if point_id in vertices:
            duplicate_ids.append(point_id)
        vertices[point_id] = _parse_float_triplet(p.text, f"P id={point_id}", coordinate_order=coordinate_order)

    if duplicate_ids:
        warnings.append(f"Duplicate point ids found while parsing: {sorted(set(duplicate_ids))}")

    faces: list[tuple[str, str, str]] = []
    for idx, f in enumerate(_iter_children(faces_elem, "F"), start=1):
        faces.append(_parse_face_ids(f.text, f"F #{idx}"))

    return TinSurface(
        name=surface.attrib.get("name"),
        surface_type=surface_type,
        vertices=vertices,
        faces=faces,
        available_surfaces=summaries,
        warnings=warnings,
        coordinate_order=coordinate_order,
    )


def _triangle_area_3d(a: tuple[float, float, float], b: tuple[float, float, float], c: tuple[float, float, float]) -> float:
    ab = (b[0] - a[0], b[1] - a[1], b[2] - a[2])
    ac = (c[0] - a[0], c[1] - a[1], c[2] - a[2])
    cross = (
        ab[1] * ac[2] - ab[2] * ac[1],
        ab[2] * ac[0] - ab[0] * ac[2],
        ab[0] * ac[1] - ab[1] * ac[0],
    )
    return 0.5 * math.sqrt(cross[0] ** 2 + cross[1] ** 2 + cross[2] ** 2)


def _connectivity_hash(faces: list[tuple[str, str, str]]) -> str:
    digest = hashlib.sha256()
    for face in faces:
        digest.update((",".join(face) + "\n").encode("utf-8"))
    return digest.hexdigest()


def _build_indices(vertices: Mapping[str, tuple[float, float, float]]) -> tuple[dict[str, int], dict[str, int]]:
    point_ids = list(vertices.keys())
    obj_indices = {pid: i + 1 for i, pid in enumerate(point_ids)}
    ply_indices = {pid: i for i, pid in enumerate(point_ids)}
    return obj_indices, ply_indices


def _validate_mesh(surface: TinSurface, expected: Mapping[str, Any], options: Mapping[str, Any]) -> dict[str, Any]:
    vertices = surface.vertices
    faces = surface.faces
    missing_refs = sorted({pid for face in faces for pid in face if pid not in vertices})

    tol = float(options.get("degenerate_area_tolerance", 1e-12))
    degenerate_faces = 0
    min_area: float | None = None
    for face in faces:
        if any(pid not in vertices for pid in face):
            continue
        area = _triangle_area_3d(vertices[face[0]], vertices[face[1]], vertices[face[2]])
        min_area = area if min_area is None else min(min_area, area)
        if area <= tol:
            degenerate_faces += 1

    coords = list(vertices.values())
    xs = [p[0] for p in coords]
    ys = [p[1] for p in coords]
    zs = [p[2] for p in coords]

    expected_points = expected.get("n_points")
    expected_faces = expected.get("n_faces")
    expected_type = expected.get("surface_type")

    checks: dict[str, bool] = {
        "missing_point_references_ok": len(missing_refs) == 0,
        "degenerate_faces_ok": degenerate_faces == 0,
    }
    if expected_points is not None:
        checks["expected_n_points_ok"] = int(expected_points) == len(vertices)
    if expected_faces is not None:
        checks["expected_n_faces_ok"] = int(expected_faces) == len(faces)
    if expected_type:
        checks["expected_surface_type_ok"] = str(expected_type).upper() == str(surface.surface_type).upper()

    validation_passed = all(checks.values())

    return {
        "surface_name": surface.name,
        "surface_type": surface.surface_type,
        "coordinate_order": surface.coordinate_order,
        "n_points_xml": len(vertices),
        "n_faces_xml": len(faces),
        "missing_point_references": len(missing_refs),
        "missing_point_reference_ids": ";".join(missing_refs),
        "degenerate_faces": degenerate_faces,
        "min_triangle_area_3d": min_area,
        "min_x": min(xs) if xs else None,
        "max_x": max(xs) if xs else None,
        "min_y": min(ys) if ys else None,
        "max_y": max(ys) if ys else None,
        "min_z": min(zs) if zs else None,
        "max_z": max(zs) if zs else None,
        "connectivity_hash": _connectivity_hash(faces),
        "validation_passed": validation_passed,
        **checks,
    }


def resolve_landxml_tin_config(raw_config: dict[str, Any]) -> dict[str, Any]:
    section = raw_config.get("landxml_tin_to_mesh") or {}
    if not isinstance(section, dict) or not section:
        raise ValueError("Missing or invalid section: landxml_tin_to_mesh")

    config = _copy_defaults(DEFAULT_CONFIG)
    _deep_update(config, section)

    name = config.get("name") or raw_config.get("name") or "landxml_tin"
    outdir = config.get("outdir") or raw_config.get("outdir") or f"outputs/{name}"
    excel = config.get("excel", raw_config.get("excel", True))
    coordinate_order = _validate_coordinate_order(str(config.get("coordinate_order", "xyz")))

    input_xml = config.get("input_xml")
    if not input_xml:
        raise ValueError("landxml_tin_to_mesh.input_xml is required")

    config["name"] = str(name)
    config["outdir"] = str(Path(str(outdir)).expanduser())
    config["excel"] = bool(excel)
    config["input_xml"] = str(Path(str(input_xml)).expanduser())
    config["coordinate_order"] = coordinate_order
    config["pipeline"] = "landxml_tin_to_mesh"

    xml_path = Path(config["input_xml"]).expanduser()
    if not xml_path.exists():
        raise FileNotFoundError(f"LandXML input not found: {xml_path}")

    if not isinstance(config.get("outputs"), dict):
        raise ValueError("landxml_tin_to_mesh.outputs must be a mapping")
    if not isinstance(config.get("options"), dict):
        raise ValueError("landxml_tin_to_mesh.options must be a mapping")
    if not isinstance(config.get("expected"), dict):
        raise ValueError("landxml_tin_to_mesh.expected must be a mapping")

    return config


def _write_vertices_csv(path: Path, vertices: Mapping[str, tuple[float, float, float]], obj_indices: Mapping[str, int], ply_indices: Mapping[str, int]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["point_id", "x", "y", "z", "obj_index", "ply_index"])
        writer.writeheader()
        for pid, (x, y, z) in vertices.items():
            writer.writerow({"point_id": pid, "x": x, "y": y, "z": z, "obj_index": obj_indices[pid], "ply_index": ply_indices[pid]})


def _write_faces_csv(path: Path, faces: list[tuple[str, str, str]], obj_indices: Mapping[str, int], ply_indices: Mapping[str, int]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        fieldnames = ["face_id", "p1", "p2", "p3", "obj_i", "obj_j", "obj_k", "ply_i", "ply_j", "ply_k"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for face_id, (p1, p2, p3) in enumerate(faces, start=1):
            writer.writerow({
                "face_id": face_id,
                "p1": p1,
                "p2": p2,
                "p3": p3,
                "obj_i": obj_indices[p1],
                "obj_j": obj_indices[p2],
                "obj_k": obj_indices[p3],
                "ply_i": ply_indices[p1],
                "ply_j": ply_indices[p2],
                "ply_k": ply_indices[p3],
            })


def _write_obj(path: Path, surface: TinSurface, obj_indices: Mapping[str, int]) -> None:
    safe_name = str(surface.name or "landxml_tin").replace(" ", "_")
    with path.open("w", encoding="utf-8") as handle:
        handle.write("# Generated from LandXML by spatial_analysis\n")
        handle.write(f"# coordinate_order={surface.coordinate_order}\n")
        handle.write(f"o {safe_name}\n")
        for x, y, z in surface.vertices.values():
            handle.write(f"v {x:.12g} {y:.12g} {z:.12g}\n")
        for p1, p2, p3 in surface.faces:
            handle.write(f"f {obj_indices[p1]} {obj_indices[p2]} {obj_indices[p3]}\n")


def _write_ply(path: Path, surface: TinSurface, ply_indices: Mapping[str, int]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        handle.write("ply\n")
        handle.write("format ascii 1.0\n")
        handle.write("comment Generated from LandXML by spatial_analysis\n")
        handle.write(f"comment coordinate_order={surface.coordinate_order}\n")
        handle.write(f"element vertex {len(surface.vertices)}\n")
        handle.write("property double x\n")
        handle.write("property double y\n")
        handle.write("property double z\n")
        handle.write(f"element face {len(surface.faces)}\n")
        handle.write("property list uchar int vertex_indices\n")
        handle.write("end_header\n")
        for x, y, z in surface.vertices.values():
            handle.write(f"{x:.12g} {y:.12g} {z:.12g}\n")
        for p1, p2, p3 in surface.faces:
            handle.write(f"3 {ply_indices[p1]} {ply_indices[p2]} {ply_indices[p3]}\n")


def _count_obj(path: Path) -> tuple[int, int]:
    v_count = 0
    f_count = 0
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if line.startswith("v "):
                v_count += 1
            elif line.startswith("f "):
                f_count += 1
    return v_count, f_count


def _write_excel_report(path: Path, config: Mapping[str, Any], metrics: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    summary_df = pd.DataFrame([metrics.get("summary", {})])
    config_df = pd.DataFrame([{"key": k, "value": v} for k, v in _flatten_mapping(config)])
    surfaces_df = pd.DataFrame(metrics.get("surfaces", []))
    validation_df = pd.DataFrame([metrics.get("validation", {})])
    file_df = pd.DataFrame(metrics.get("file_inventory", []))

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        config_df.to_excel(writer, sheet_name="Config", index=False)
        surfaces_df.to_excel(writer, sheet_name="Surfaces", index=False)
        validation_df.to_excel(writer, sheet_name="MeshValidation", index=False)
        file_df.to_excel(writer, sheet_name="FileInventory", index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 70)
    wb.save(path)


def run_landxml_tin_to_mesh(config: dict[str, Any]) -> dict[str, str]:
    xml_path = Path(config["input_xml"]).expanduser().resolve()
    outdir = Path(config["outdir"]).expanduser().resolve()
    mesh_dir = outdir / "mesh"
    report_dir = outdir / "report"
    metadata_dir = outdir / "metadata"
    mesh_dir.mkdir(parents=True, exist_ok=True)
    report_dir.mkdir(parents=True, exist_ok=True)
    metadata_dir.mkdir(parents=True, exist_ok=True)

    outputs_cfg = config["outputs"]
    options = config["options"]
    expected = config["expected"]

    surface = extract_tin_surface(xml_path, config.get("surface_name"), coordinate_order=config.get("coordinate_order", "xyz"))
    validation = _validate_mesh(surface, expected, options)

    strict_faces = bool(options.get("strict_faces", True))
    if strict_faces and validation["missing_point_references"]:
        raise ValueError(f"Faces reference missing point ids: {validation['missing_point_reference_ids']}")

    validate_counts = bool(options.get("validate_counts", True))
    if validate_counts:
        failed_expected = [k for k in ("expected_n_points_ok", "expected_n_faces_ok", "expected_surface_type_ok") if validation.get(k) is False]
        if failed_expected:
            raise ValueError(f"LandXML TIN expected-count/type validation failed: {', '.join(failed_expected)}")

    obj_indices, ply_indices = _build_indices(surface.vertices)

    vertices_path = mesh_dir / str(outputs_cfg.get("vertices_csv", "vertices.csv"))
    faces_path = mesh_dir / str(outputs_cfg.get("faces_csv", "faces.csv"))
    obj_path = mesh_dir / str(outputs_cfg.get("obj", "tin_mesh.obj"))
    ply_path = mesh_dir / str(outputs_cfg.get("ply", "tin_mesh.ply"))
    excel_path = report_dir / str(outputs_cfg.get("excel_report", "landxml_tin_report.xlsx"))
    summary_json_path = metadata_dir / str(outputs_cfg.get("summary_json", "landxml_tin_summary.json"))

    _write_vertices_csv(vertices_path, surface.vertices, obj_indices, ply_indices)
    _write_faces_csv(faces_path, surface.faces, obj_indices, ply_indices)
    _write_obj(obj_path, surface, obj_indices)

    file_inventory: list[dict[str, str]] = [
        {"label": "vertices_csv", "path": str(vertices_path)},
        {"label": "faces_csv", "path": str(faces_path)},
        {"label": "obj", "path": str(obj_path)},
    ]

    if bool(options.get("write_ply", True)):
        _write_ply(ply_path, surface, ply_indices)
        file_inventory.append({"label": "ply", "path": str(ply_path)})

    obj_v_count, obj_f_count = _count_obj(obj_path)
    validation["obj_vertex_count"] = obj_v_count
    validation["obj_face_count"] = obj_f_count
    validation["obj_counts_ok"] = obj_v_count == len(surface.vertices) and obj_f_count == len(surface.faces)
    validation["validation_passed"] = bool(validation.get("validation_passed")) and bool(validation["obj_counts_ok"])

    metrics: dict[str, Any] = {
        "summary": {
            "timestamp": datetime.now().astimezone().isoformat(),
            "pipeline": "landxml_tin_to_mesh",
            "input_xml": str(xml_path),
            "selected_surface": surface.name,
            "surface_type": surface.surface_type,
            "coordinate_order": surface.coordinate_order,
            "n_points": len(surface.vertices),
            "n_faces": len(surface.faces),
            "validation_passed": validation["validation_passed"],
            "warnings": "; ".join(surface.warnings),
        },
        "surfaces": surface.available_surfaces,
        "validation": validation,
        "file_inventory": file_inventory,
    }

    if config.get("excel", True):
        _write_excel_report(excel_path, config, metrics)
        file_inventory.append({"label": "excel_report", "path": str(excel_path)})

    summary_json_path.write_text(json.dumps(metrics, indent=2), encoding="utf-8")
    file_inventory.append({"label": "summary_json", "path": str(summary_json_path)})

    return {item["label"]: item["path"] for item in file_inventory}
