from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, MutableMapping, Tuple

import importlib
import importlib.util
import json

import numpy as np
import pandas as pd
import rasterio
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from rasterio.features import rasterize
from rasterio.warp import transform_geom

from raster_compare.core import DEFAULT_NODATA


_HAS_FIONA = importlib.util.find_spec("fiona") is not None
_HAS_SCIPY = importlib.util.find_spec("scipy") is not None


DEFAULT_CONFIG: Dict[str, Any] = {
    "excel": True,
    "outputs": {
        "adapted_raster": "adapted_raster.tif",
        "excel_report": "raster_adapt_polygon_report.xlsx",
        "summary_json": "raster_adapt_polygon_summary.json",
        "save_intermediates": True,
    },
    "reference_ring": {
        "outer_buffer_px": 20,
        "inner_buffer_px": 0,
        "min_reference_pixels": 500,
    },
    "adaptation": {
        "method": "boundary_idw",
        "idw_power": 2.0,
        "k_nearest": 32,
        "max_reference_points": 10000,
        "random_seed": 42,
    },
    "border_blending": {
        "enabled": True,
        "blend_width_px": 5,
    },
    "nodata": {
        "preserve_nodata": True,
    },
}


@dataclass
class AdaptOutputs:
    adapted_raster: Path
    report_path: Path | None
    summary_json: Path
    adapted_surface: Path | None
    modify_mask: Path | None
    reference_ring: Path | None
    blend_weights: Path | None


def _deep_update(base: MutableMapping[str, Any], updates: Mapping[str, Any]) -> MutableMapping[str, Any]:
    for key, value in updates.items():
        if isinstance(value, Mapping) and isinstance(base.get(key), MutableMapping):
            base[key] = _deep_update(base[key], value)  # type: ignore[index]
        else:
            base[key] = value
    return base


def _copy_defaults(data: Mapping[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for key, value in data.items():
        if isinstance(value, Mapping):
            out[key] = _copy_defaults(value)
        else:
            out[key] = value
    return out


def _flatten_mapping(data: Mapping[str, Any], prefix: str = "") -> Iterable[Tuple[str, Any]]:
    for key, value in data.items():
        full_key = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, Mapping):
            yield from _flatten_mapping(value, full_key)
        else:
            yield full_key, value


def _validate_path(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{label} not found: {path}")


def _load_polygon_geometries(path: Path, target_crs: rasterio.crs.CRS) -> List[Dict[str, Any]]:
    if not _HAS_FIONA:
        raise ImportError("fiona is required to read polygon inputs. Install it via conda or pip.")

    fiona = importlib.import_module("fiona")
    geoms: List[Dict[str, Any]] = []
    with fiona.open(path) as src:
        src_crs = rasterio.crs.CRS.from_user_input(src.crs) if src.crs else None
        for feature in src:
            geom = feature.get("geometry")
            if geom is None:
                continue
            if src_crs and src_crs != target_crs:
                geom = transform_geom(src_crs, target_crs, geom)
            geoms.append(geom)

    if not geoms:
        raise ValueError(f"Polygon file contains no geometries: {path}")
    return geoms


def _rasterize_polygon(path: Path, ref_ds: rasterio.io.DatasetReader) -> np.ndarray:
    geoms = _load_polygon_geometries(path, ref_ds.crs)
    mask = rasterize(
        [(geom, 1) for geom in geoms],
        out_shape=(ref_ds.height, ref_ds.width),
        transform=ref_ds.transform,
        fill=0,
        dtype="uint8",
    )
    return mask.astype(bool)


def _mask_from_nodata(data: np.ndarray, nodata: float | None) -> np.ndarray:
    mask = ~np.isfinite(data)
    if nodata is not None:
        mask |= data == nodata
    return mask


def _distance_inside(mask: np.ndarray) -> np.ndarray:
    if not _HAS_SCIPY:
        raise ImportError("scipy is required for raster_adapt_polygon. Install scipy in the environment.")
    ndimage = importlib.import_module("scipy.ndimage")
    return ndimage.distance_transform_edt(mask)


def _make_reference_ring(modify_mask: np.ndarray, outer_buffer_px: int, inner_buffer_px: int) -> np.ndarray:
    if not _HAS_SCIPY:
        raise ImportError("scipy is required for raster_adapt_polygon. Install scipy in the environment.")
    ndimage = importlib.import_module("scipy.ndimage")

    outer = ndimage.binary_dilation(modify_mask, iterations=int(outer_buffer_px))
    if inner_buffer_px > 0:
        inner = ndimage.binary_dilation(modify_mask, iterations=int(inner_buffer_px))
    else:
        inner = modify_mask
    return outer & ~inner


def _pixel_centers(transform: rasterio.Affine, rows: np.ndarray, cols: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    xs = transform.c + (cols + 0.5) * transform.a + (rows + 0.5) * transform.b
    ys = transform.f + (cols + 0.5) * transform.d + (rows + 0.5) * transform.e
    return xs.astype(np.float64), ys.astype(np.float64)


def _subsample_reference_points(
    coords: np.ndarray,
    values: np.ndarray,
    max_reference_points: int | None,
    random_seed: int | None,
) -> Tuple[np.ndarray, np.ndarray, int]:
    n = coords.shape[0]
    if max_reference_points is None or max_reference_points <= 0 or n <= max_reference_points:
        return coords, values, n
    rng = np.random.default_rng(random_seed)
    idx = rng.choice(n, size=int(max_reference_points), replace=False)
    return coords[idx], values[idx], int(max_reference_points)


def _boundary_idw(
    target_coords: np.ndarray,
    reference_coords: np.ndarray,
    reference_values: np.ndarray,
    idw_power: float,
    k_nearest: int,
) -> np.ndarray:
    if reference_coords.size == 0:
        raise ValueError("No reference points available for IDW interpolation.")
    if not _HAS_SCIPY:
        raise ImportError("scipy is required for boundary_idw interpolation.")

    spatial = importlib.import_module("scipy.spatial")
    tree = spatial.cKDTree(reference_coords)
    k = min(int(k_nearest), reference_coords.shape[0])
    distances, indices = tree.query(target_coords, k=k)

    if k == 1:
        distances = distances[:, None]
        indices = indices[:, None]

    values = reference_values[indices]
    exact = distances <= 1e-12
    weights = np.zeros_like(distances, dtype=np.float64)
    non_exact = ~exact
    weights[non_exact] = 1.0 / np.power(distances[non_exact], float(idw_power))

    has_exact = exact.any(axis=1)
    out = np.empty(target_coords.shape[0], dtype=np.float32)
    if has_exact.any():
        first_exact = np.argmax(exact[has_exact], axis=1)
        out[has_exact] = values[has_exact, first_exact].astype(np.float32)
    if (~has_exact).any():
        w = weights[~has_exact]
        v = values[~has_exact]
        out[~has_exact] = (np.sum(w * v, axis=1) / np.sum(w, axis=1)).astype(np.float32)
    return out


def _write_raster(path: Path, data: np.ndarray, profile: rasterio.profiles.Profile) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with rasterio.open(path, "w", **profile) as dst:
        dst.write(data.astype(np.float32), 1)


def _stats(values: np.ndarray) -> Dict[str, Any]:
    values = values[np.isfinite(values)]
    if values.size == 0:
        return {"count": 0, "min": None, "max": None, "mean": None, "median": None, "std": None}
    return {
        "count": int(values.size),
        "min": float(np.min(values)),
        "max": float(np.max(values)),
        "mean": float(np.mean(values)),
        "median": float(np.median(values)),
        "std": float(np.std(values)),
    }


def resolve_raster_adapt_polygon_config(raw_config: Mapping[str, Any]) -> Dict[str, Any]:
    section = raw_config.get("raster_adapt_polygon") or {}
    if not isinstance(section, Mapping) or not section:
        raise ValueError("Missing or invalid section: raster_adapt_polygon")

    config = _copy_defaults(DEFAULT_CONFIG)
    _deep_update(config, section)

    name = config.get("name") or raw_config.get("name") or "raster_adapt_polygon"
    outdir = config.get("outdir") or raw_config.get("outdir") or f"outputs/{name}"
    excel = config.get("excel", raw_config.get("excel", True))

    if not config.get("raster"):
        raise ValueError("raster_adapt_polygon.raster is required")
    if not config.get("modify_polygon"):
        raise ValueError("raster_adapt_polygon.modify_polygon is required")

    config["name"] = str(name)
    config["outdir"] = str(Path(str(outdir)).expanduser())
    config["excel"] = bool(excel)
    config["raster"] = str(Path(str(config["raster"])).expanduser())
    config["modify_polygon"] = str(Path(str(config["modify_polygon"])).expanduser())
    config["pipeline"] = "raster_adapt_polygon"

    _validate_path(Path(config["raster"]), "raster")
    _validate_path(Path(config["modify_polygon"]), "modify_polygon")

    ring = config["reference_ring"]
    ring["outer_buffer_px"] = int(ring.get("outer_buffer_px", 20))
    ring["inner_buffer_px"] = int(ring.get("inner_buffer_px", 0))
    ring["min_reference_pixels"] = int(ring.get("min_reference_pixels", 500))
    if ring["outer_buffer_px"] <= ring["inner_buffer_px"]:
        raise ValueError("reference_ring.outer_buffer_px must be greater than inner_buffer_px")
    if ring["inner_buffer_px"] < 0 or ring["min_reference_pixels"] <= 0:
        raise ValueError("reference_ring inner buffer must be >= 0 and min_reference_pixels must be > 0")

    adaptation = config["adaptation"]
    method = str(adaptation.get("method", "boundary_idw")).lower().strip()
    if method != "boundary_idw":
        raise ValueError("Only adaptation.method='boundary_idw' is supported in V1")
    adaptation["method"] = method
    adaptation["idw_power"] = float(adaptation.get("idw_power", 2.0))
    adaptation["k_nearest"] = int(adaptation.get("k_nearest", 32))
    adaptation["max_reference_points"] = int(adaptation.get("max_reference_points", 10000))
    random_seed = adaptation.get("random_seed", 42)
    adaptation["random_seed"] = None if random_seed is None else int(random_seed)
    if adaptation["idw_power"] <= 0:
        raise ValueError("adaptation.idw_power must be > 0")
    if adaptation["k_nearest"] <= 0:
        raise ValueError("adaptation.k_nearest must be > 0")

    blending = config["border_blending"]
    blending["enabled"] = bool(blending.get("enabled", True))
    blending["blend_width_px"] = int(blending.get("blend_width_px", 5))
    if blending["blend_width_px"] < 0:
        raise ValueError("border_blending.blend_width_px must be >= 0")

    config["outputs"]["save_intermediates"] = bool(config["outputs"].get("save_intermediates", True))
    config["nodata"]["preserve_nodata"] = bool(config["nodata"].get("preserve_nodata", True))
    return config


def _write_excel_report(report_path: Path, config: Mapping[str, Any], metrics: Mapping[str, Any]) -> Path:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        pd.DataFrame([metrics.get("summary", {})]).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame([{"key": k, "value": v} for k, v in _flatten_mapping(config)]).to_excel(writer, sheet_name="Config", index=False)
        pd.DataFrame([metrics.get("reference_ring", {})]).to_excel(writer, sheet_name="ReferenceRing", index=False)
        pd.DataFrame([metrics.get("interpolation", {})]).to_excel(writer, sheet_name="Interpolation", index=False)
        pd.DataFrame(metrics.get("raster_stats", [])).to_excel(writer, sheet_name="RasterStats", index=False)
        pd.DataFrame(metrics.get("file_inventory", [])).to_excel(writer, sheet_name="FileInventory", index=False)

    wb = load_workbook(report_path)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 70)
    wb.save(report_path)
    return report_path


def run_raster_adapt_polygon(config: Mapping[str, Any]) -> Dict[str, str]:
    raster_path = Path(config["raster"]).expanduser().resolve()
    polygon_path = Path(config["modify_polygon"]).expanduser().resolve()
    outdir = Path(config["outdir"]).expanduser().resolve()
    name = str(config["name"])

    rasters_dir = outdir / "rasters"
    report_dir = outdir / "report"
    metadata_dir = outdir / "metadata"
    rasters_dir.mkdir(parents=True, exist_ok=True)
    report_dir.mkdir(parents=True, exist_ok=True)
    metadata_dir.mkdir(parents=True, exist_ok=True)

    outputs_cfg = config["outputs"]
    save_intermediates = bool(outputs_cfg.get("save_intermediates", True))

    adapted_raster_path = rasters_dir / str(outputs_cfg.get("adapted_raster", "adapted_raster.tif"))
    adapted_surface_path = rasters_dir / f"{name}_adapted_surface.tif"
    modify_mask_path = rasters_dir / f"{name}_modify_mask.tif"
    reference_ring_path = rasters_dir / f"{name}_reference_ring.tif"
    blend_weights_path = rasters_dir / f"{name}_blend_weights.tif"
    report_path = report_dir / str(outputs_cfg.get("excel_report", "raster_adapt_polygon_report.xlsx"))
    summary_json_path = metadata_dir / str(outputs_cfg.get("summary_json", "raster_adapt_polygon_summary.json"))

    with rasterio.open(raster_path) as src:
        if src.crs is None:
            raise ValueError("Input raster must have a CRS to rasterize/reproject the polygon correctly.")
        data = src.read(1).astype(np.float32)
        profile = src.profile.copy()
        nodata = src.nodata if src.nodata is not None else DEFAULT_NODATA
        valid_mask = ~_mask_from_nodata(data, src.nodata)
        modify_mask = _rasterize_polygon(polygon_path, src)
        transform = src.transform

    inside_valid = modify_mask & valid_mask
    if int(np.sum(modify_mask)) == 0:
        raise ValueError("modify_polygon does not overlap the raster grid.")
    if int(np.sum(inside_valid)) == 0:
        raise ValueError("No valid raster pixels found inside modify_polygon.")

    ring_cfg = config["reference_ring"]
    ring_mask = _make_reference_ring(modify_mask, ring_cfg["outer_buffer_px"], ring_cfg["inner_buffer_px"])
    reference_mask = ring_mask & valid_mask
    n_reference = int(np.sum(reference_mask))
    if n_reference < int(ring_cfg["min_reference_pixels"]):
        raise ValueError(
            f"Not enough valid reference pixels: {n_reference} < {ring_cfg['min_reference_pixels']}. "
            "Increase outer_buffer_px or lower min_reference_pixels."
        )

    ref_rows, ref_cols = np.where(reference_mask)
    ref_x, ref_y = _pixel_centers(transform, ref_rows, ref_cols)
    ref_coords = np.column_stack([ref_x, ref_y]).astype(np.float64)
    ref_values = data[ref_rows, ref_cols].astype(np.float64)

    adaptation_cfg = config["adaptation"]
    ref_coords_used, ref_values_used, n_reference_used = _subsample_reference_points(
        ref_coords,
        ref_values,
        adaptation_cfg.get("max_reference_points"),
        adaptation_cfg.get("random_seed"),
    )

    target_rows, target_cols = np.where(modify_mask)
    target_x, target_y = _pixel_centers(transform, target_rows, target_cols)
    target_coords = np.column_stack([target_x, target_y]).astype(np.float64)

    interpolated = _boundary_idw(
        target_coords=target_coords,
        reference_coords=ref_coords_used,
        reference_values=ref_values_used,
        idw_power=float(adaptation_cfg["idw_power"]),
        k_nearest=int(adaptation_cfg["k_nearest"]),
    )

    adapted_surface = np.full(data.shape, float(nodata), dtype=np.float32)
    adapted_surface[target_rows, target_cols] = interpolated

    blending_cfg = config["border_blending"]
    if bool(blending_cfg.get("enabled", True)) and int(blending_cfg.get("blend_width_px", 0)) > 0:
        dist_inside = _distance_inside(modify_mask)
        weights = np.clip(dist_inside / float(blending_cfg["blend_width_px"]), 0.0, 1.0).astype(np.float32)
        weights[~modify_mask] = 0.0
    else:
        weights = modify_mask.astype(np.float32)

    adapted = data.copy()
    adapted_values = data[modify_mask] * (1.0 - weights[modify_mask]) + adapted_surface[modify_mask] * weights[modify_mask]
    adapted[modify_mask] = adapted_values.astype(np.float32)

    preserve_nodata = bool(config.get("nodata", {}).get("preserve_nodata", True))
    if preserve_nodata:
        adapted[~valid_mask] = float(nodata)
        adapted_surface[~valid_mask & modify_mask] = float(nodata)

    out_profile = profile.copy()
    out_profile.update({"driver": "GTiff", "dtype": "float32", "count": 1, "nodata": nodata, "compress": "deflate"})
    _write_raster(adapted_raster_path, adapted, out_profile)

    file_inventory: List[Dict[str, str]] = [{"label": "adapted_raster", "path": str(adapted_raster_path)}]
    if save_intermediates:
        _write_raster(adapted_surface_path, adapted_surface, out_profile)
        mask_profile = out_profile.copy()
        mask_profile.update({"dtype": "float32", "nodata": 0.0})
        _write_raster(modify_mask_path, modify_mask.astype(np.float32), mask_profile)
        _write_raster(reference_ring_path, reference_mask.astype(np.float32), mask_profile)
        _write_raster(blend_weights_path, weights.astype(np.float32), mask_profile)
        file_inventory.extend(
            [
                {"label": "adapted_surface", "path": str(adapted_surface_path)},
                {"label": "modify_mask", "path": str(modify_mask_path)},
                {"label": "reference_ring", "path": str(reference_ring_path)},
                {"label": "blend_weights", "path": str(blend_weights_path)},
            ]
        )

    original_inside = data[inside_valid]
    adapted_inside = adapted[inside_valid]
    diff_inside = adapted_inside - original_inside

    metrics: Dict[str, Any] = {
        "summary": {
            "timestamp": datetime.now().astimezone().isoformat(),
            "pipeline": "raster_adapt_polygon",
            "name": name,
            "raster": str(raster_path),
            "modify_polygon": str(polygon_path),
            "method": adaptation_cfg["method"],
            "modified_pixels": int(np.sum(modify_mask)),
            "valid_modified_pixels": int(np.sum(inside_valid)),
            "reference_pixels": n_reference,
            "reference_pixels_used": n_reference_used,
            "adapted_raster": str(adapted_raster_path),
        },
        "reference_ring": {
            "outer_buffer_px": ring_cfg["outer_buffer_px"],
            "inner_buffer_px": ring_cfg["inner_buffer_px"],
            "min_reference_pixels": ring_cfg["min_reference_pixels"],
            "reference_pixels": n_reference,
            **{f"reference_z_{k}": v for k, v in _stats(ref_values).items()},
        },
        "interpolation": {
            "method": adaptation_cfg["method"],
            "idw_power": adaptation_cfg["idw_power"],
            "k_nearest": adaptation_cfg["k_nearest"],
            "max_reference_points": adaptation_cfg["max_reference_points"],
            "reference_points_used": n_reference_used,
            "target_pixels_interpolated": int(target_coords.shape[0]),
        },
        "raster_stats": [
            {"scope": "original_inside_polygon", **_stats(original_inside)},
            {"scope": "adapted_inside_polygon", **_stats(adapted_inside)},
            {"scope": "adapted_minus_original_inside_polygon", **_stats(diff_inside)},
        ],
        "file_inventory": file_inventory,
    }

    if config.get("excel", True):
        _write_excel_report(report_path, config, metrics)
        file_inventory.append({"label": "excel_report", "path": str(report_path)})

    summary_json_path.write_text(json.dumps(metrics, indent=2), encoding="utf-8")
    file_inventory.append({"label": "summary_json", "path": str(summary_json_path)})

    return {item["label"]: item["path"] for item in file_inventory}
