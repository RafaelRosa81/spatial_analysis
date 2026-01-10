from __future__ import annotations

from datetime import datetime
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Tuple

import numpy as np
import pandas as pd
import rasterio
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font


def load_valid_values(raster_path: Path) -> np.ndarray:
    """
    Read band 1 and return a 1D array of valid values (nodata/NaN removed).
    """
    raster_path = Path(raster_path)
    with rasterio.open(raster_path) as src:
        arr = src.read(1)
        nodata = src.nodata

    v = arr.reshape(-1)

    # Remove NaNs
    v = v[np.isfinite(v)]

    # Remove nodata if present
    if nodata is not None:
        v = v[v != nodata]

    return v.astype(np.float64, copy=False)


def compute_stats(values: np.ndarray) -> Dict[str, float]:
    """
    Compute descriptive statistics + RMSE and selected percentiles.
    """
    keys = ["min", "max", "mean", "median", "std", "rmse", "p01", "p05", "p25", "p50", "p75", "p95", "p99"]
    if values.size == 0:
        return {k: float("nan") for k in keys}

    v = values.astype(np.float64, copy=False)
    rmse = float(np.sqrt(np.mean(v * v)))

    return {
        "min": float(np.min(v)),
        "max": float(np.max(v)),
        "mean": float(np.mean(v)),
        "median": float(np.median(v)),
        "std": float(np.std(v, ddof=0)),
        "rmse": rmse,
        "p01": float(np.quantile(v, 0.01)),
        "p05": float(np.quantile(v, 0.05)),
        "p25": float(np.quantile(v, 0.25)),
        "p50": float(np.quantile(v, 0.50)),
        "p75": float(np.quantile(v, 0.75)),
        "p95": float(np.quantile(v, 0.95)),
        "p99": float(np.quantile(v, 0.99)),
    }


def threshold_table(abs_values: np.ndarray, thresholds: List[float]) -> pd.DataFrame:
    """
    Build bins: <=t1, (t1,t2], ..., >t_last
    """
    thresholds = sorted([float(t) for t in thresholds])
    edges = [0.0] + thresholds + [np.inf]

    labels: List[str] = []
    for i in range(len(edges) - 1):
        lo, hi = edges[i], edges[i + 1]
        if hi == np.inf:
            labels.append(f"> {lo:g}")
        elif lo == 0.0:
            labels.append(f"<= {hi:g}")
        else:
            labels.append(f"({lo:g}, {hi:g}]")

    counts: List[int] = []
    for i in range(len(edges) - 1):
        lo, hi = edges[i], edges[i + 1]
        if hi == np.inf:
            c = int(np.sum(abs_values > lo))
        elif lo == 0.0:
            c = int(np.sum(abs_values <= hi))
        else:
            c = int(np.sum((abs_values > lo) & (abs_values <= hi)))
        counts.append(c)

    total = int(abs_values.size)
    perc = [(c / total * 100.0) if total > 0 else float("nan") for c in counts]

    return pd.DataFrame({"Bin": labels, "Count": counts, "Percent": perc})


def histogram_table(values: np.ndarray, bins: int = 60) -> pd.DataFrame:
    if values.size == 0:
        return pd.DataFrame({"bin_left": [], "bin_right": [], "count": [], "percent": []})

    hist, edges = np.histogram(values, bins=int(bins))
    total = int(hist.sum())

    return pd.DataFrame(
        {
            "bin_left": edges[:-1],
            "bin_right": edges[1:],
            "count": hist.astype(int),
            "percent": (hist / total * 100.0) if total > 0 else np.nan,
        }
    )


def _meta(path: Path) -> Dict[str, str]:
    with rasterio.open(path) as src:
        return {
            "path": str(path),
            "crs": str(src.crs),
            "width": str(src.width),
            "height": str(src.height),
            "pixel_x": str(src.transform.a),
            "pixel_y": str(abs(src.transform.e)),
            "nodata": str(src.nodata),
            "dtype": str(src.dtypes[0]),
            "bounds": str(src.bounds),
        }


def _crs_uses_meters(crs: rasterio.crs.CRS | None) -> bool:
    if crs is None:
        return False
    if crs.is_projected:
        return True
    return False


def _pixel_area_m2(raster_path: Path) -> Tuple[float | None, str | None]:
    with rasterio.open(raster_path) as src:
        crs = src.crs
        transform = src.transform

    if _crs_uses_meters(crs):
        return abs(transform.a * transform.e), None
    if crs is None:
        return None, "Pixel area not computed: raster has no CRS."
    if not crs.is_projected:
        return None, "Pixel area not computed: CRS is not projected."
    return None, "Pixel area not computed: CRS units are not meters."


def _crs_string(crs: rasterio.crs.CRS | None) -> str | None:
    if crs is None:
        return None
    epsg = crs.to_epsg()
    if epsg is not None:
        return f"EPSG:{epsg}"
    try:
        return crs.to_wkt()
    except Exception:
        return str(crs)


def collect_raster_metadata(path: Path) -> Dict[str, Any]:
    warnings: List[str] = []
    path = Path(path)
    with rasterio.open(path) as src:
        crs = src.crs
        transform = src.transform
        bounds = src.bounds

        if crs is None:
            warnings.append("CRS missing; some alignment comparisons may be unavailable.")

        if transform is None:
            warnings.append("Transform missing; grid shift metrics may be unavailable.")

        meta: Dict[str, Any] = {
            "path": str(path),
            "crs": _crs_string(crs),
            "crs_wkt": crs.to_wkt() if crs is not None else None,
            "is_projected": bool(crs.is_projected) if crs is not None else None,
            "width": int(src.width),
            "height": int(src.height),
            "pixel_x": float(transform.a) if transform is not None else None,
            "pixel_y": float(abs(transform.e)) if transform is not None else None,
            "nodata": src.nodata,
            "dtype": str(src.dtypes[0]),
            "bounds": {
                "left": float(bounds.left),
                "bottom": float(bounds.bottom),
                "right": float(bounds.right),
                "top": float(bounds.top),
            }
            if bounds is not None
            else None,
            "transform": {
                "a": float(transform.a),
                "b": float(transform.b),
                "c": float(transform.c),
                "d": float(transform.d),
                "e": float(transform.e),
                "f": float(transform.f),
            }
            if transform is not None
            else None,
        }
        if warnings:
            meta["warnings"] = warnings
        return meta


def _is_close(a: float | None, b: float | None) -> bool:
    if a is None or b is None:
        return False
    return np.isclose(a, b)


def _pct_change(old: float | None, new: float | None) -> float | None:
    if old is None or new is None or old == 0:
        return None
    return (new - old) / old * 100.0


def diff_alignment(ref_meta: Mapping[str, Any], src_meta: Mapping[str, Any], out_meta: Mapping[str, Any]) -> Dict[str, Any]:
    warnings: List[str] = []

    src_crs = src_meta.get("crs")
    out_crs = out_meta.get("crs")
    crs_changed = src_crs != out_crs if src_crs is not None or out_crs is not None else None

    src_px = src_meta.get("pixel_x")
    src_py = src_meta.get("pixel_y")
    out_px = out_meta.get("pixel_x")
    out_py = out_meta.get("pixel_y")

    pixel_size_changed = None
    if src_px is None or src_py is None or out_px is None or out_py is None:
        warnings.append("Pixel size missing; pixel size change metrics may be incomplete.")
    else:
        pixel_size_changed = (not _is_close(src_px, out_px)) or (not _is_close(src_py, out_py))

    src_width = src_meta.get("width")
    src_height = src_meta.get("height")
    out_width = out_meta.get("width")
    out_height = out_meta.get("height")
    shape_changed = None
    if src_width is None or src_height is None or out_width is None or out_height is None:
        warnings.append("Raster shape missing; shape change metrics may be incomplete.")
    else:
        shape_changed = (int(src_width) != int(out_width)) or (int(src_height) != int(out_height))

    ref_bounds = ref_meta.get("bounds")
    out_bounds = out_meta.get("bounds")
    bounds_changed = None
    bounds_deltas = None
    if ref_bounds is None or out_bounds is None:
        warnings.append("Bounds missing; bounds deltas may be unavailable.")
    else:
        bounds_changed = any(
            not _is_close(float(ref_bounds[key]), float(out_bounds[key]))
            for key in ("left", "bottom", "right", "top")
        )
        bounds_deltas = {
            key: float(out_bounds[key]) - float(ref_bounds[key]) for key in ("left", "bottom", "right", "top")
        }

    ref_transform = ref_meta.get("transform")
    out_transform = out_meta.get("transform")
    grid_origin_shift = None
    shift_pixels = None
    if ref_transform is None or out_transform is None:
        warnings.append("Transform missing; grid shift metrics may be unavailable.")
    else:
        dx0 = float(out_transform["c"]) - float(ref_transform["c"])
        dy0 = float(out_transform["f"]) - float(ref_transform["f"])
        grid_origin_shift = {"dx": dx0, "dy": dy0}
        if out_px in (None, 0) or out_py in (None, 0):
            warnings.append("Pixel size missing; shift-in-pixels metric may be unavailable.")
        else:
            shift_pixels = {"dx": dx0 / float(out_px), "dy": dy0 / float(out_py)}

    resampling_applied = bool(pixel_size_changed or shape_changed) if pixel_size_changed is not None or shape_changed is not None else None

    diff: Dict[str, Any] = {
        "crs_changed": crs_changed,
        "pixel_size_changed": {
            "changed": pixel_size_changed,
            "old": {"x": src_px, "y": src_py},
            "new": {"x": out_px, "y": out_py},
            "pct_change": {"x": _pct_change(src_px, out_px), "y": _pct_change(src_py, out_py)},
        },
        "shape_changed": {
            "changed": shape_changed,
            "old": {"width": src_width, "height": src_height},
            "new": {"width": out_width, "height": out_height},
        },
        "bounds_changed": {
            "changed": bounds_changed,
            "reference": ref_bounds,
            "aligned": out_bounds,
            "deltas": bounds_deltas,
        },
        "grid_origin_shift": grid_origin_shift,
        "shift_pixels": shift_pixels,
        "resampling_applied": resampling_applied,
    }

    if warnings:
        diff["warnings"] = warnings
    return diff


def build_alignment_report(
    ref_path: Path,
    src_path: Path,
    aligned_path: Path,
) -> Dict[str, Any]:
    ref_meta = collect_raster_metadata(ref_path)
    src_meta = collect_raster_metadata(src_path)
    out_meta = collect_raster_metadata(aligned_path)
    return {
        "reference": ref_meta,
        "source": src_meta,
        "aligned": out_meta,
        "diff": diff_alignment(ref_meta, src_meta, out_meta),
        "timestamp": datetime.now().astimezone().isoformat(),
    }


def _flatten_report_section(section: str, data: Mapping[str, Any]) -> Iterable[Dict[str, Any]]:
    def walk(prefix: str, value: Any) -> Iterable[Tuple[str, Any]]:
        if isinstance(value, Mapping):
            for key, inner in value.items():
                inner_prefix = f"{prefix}.{key}" if prefix else str(key)
                yield from walk(inner_prefix, inner)
        else:
            yield prefix, value

    for key, value in walk("", data):
        yield {"section": section, "field": key, "value": value}


def _report_rows(report: Mapping[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for section in ("reference", "source", "aligned", "diff"):
        data = report.get(section, {})
        if isinstance(data, Mapping):
            rows.extend(list(_flatten_report_section(section, data)))
    rows.append({"section": "report", "field": "timestamp", "value": report.get("timestamp")})
    return rows


def write_alignment_report_json_csv(
    outdir: Path,
    name: str,
    ref_path: Path,
    src_path: Path,
    aligned_path: Path,
) -> Tuple[Path, Path]:
    report_dir = Path(outdir) / "report"
    report_dir.mkdir(parents=True, exist_ok=True)

    report = build_alignment_report(ref_path=ref_path, src_path=src_path, aligned_path=aligned_path)

    json_path = report_dir / f"{name}_alignment_report.json"
    csv_path = report_dir / f"{name}_alignment_report.csv"

    json_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    pd.DataFrame(_report_rows(report)).to_csv(csv_path, index=False)

    print(f"Alignment report written: {json_path} {csv_path}")
    return json_path, csv_path


def write_excel_report(
    dz_path: Path,
    abs_dz_path: Path,
    raster1_path: Path,
    raster2_path: Path,
    out_xlsx: Path,
    thresholds: List[float],
    bins: int = 60,
    run_config: Mapping[str, Any] | None = None,
    alignment_report: Mapping[str, Any] | None = None,
) -> Path:
    """
    Create an Excel file with:
      - Stats_dz
      - Area_by_change_magnitude
      - Histogram_dz
      - Metadata
    """
    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    dz_vals = load_valid_values(dz_path)
    abs_vals = np.abs(dz_vals)

    stats_df = pd.DataFrame([compute_stats(dz_vals)])
    th_df = threshold_table(abs_vals, thresholds)
    pixel_area_m2, area_warning = _pixel_area_m2(abs_dz_path)
    if pixel_area_m2 is None:
        th_df["PixelArea_m2"] = np.nan
        th_df["Area_m2"] = np.nan
    else:
        th_df["PixelArea_m2"] = float(pixel_area_m2)
        th_df["Area_m2"] = th_df["Count"] * float(pixel_area_m2)
    hist_df = histogram_table(dz_vals, bins=bins)

    meta_rows = []
    for layer_name, p in {
        "raster1": raster1_path,
        "raster2": raster2_path,
        "dz": dz_path,
        "abs_dz": abs_dz_path,
    }.items():
        md = _meta(Path(p))
        for k, v in md.items():
            meta_rows.append({"layer": layer_name, "field": k, "value": v})
    if area_warning:
        meta_rows.append({"layer": "report", "field": "area_units_warning", "value": area_warning})
    meta_df = pd.DataFrame(meta_rows)

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        stats_df.to_excel(writer, sheet_name="Stats_dz", index=False)
        th_df.to_excel(writer, sheet_name="Area_by_change_magnitude", index=False)
        hist_df.to_excel(writer, sheet_name="Histogram_dz", index=False)
        meta_df.to_excel(writer, sheet_name="Metadata", index=False)
        if alignment_report:
            alignment_df = pd.DataFrame(_report_rows(alignment_report))
            alignment_df.to_excel(writer, sheet_name="Alignment", index=False)
        if run_config:
            config_df = pd.DataFrame(
                [{"key": k, "value": v} for k, v in run_config.items()],
                columns=["key", "value"],
            )
            config_df.to_excel(writer, sheet_name="Config", index=False)

    # Light formatting
    wb = load_workbook(out_xlsx)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = 0
            for c in col:
                if c.value is None:
                    continue
                max_len = max(max_len, len(str(c.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    if "Histogram_dz" in wb.sheetnames:
        ws = wb["Histogram_dz"]
        if ws.max_row >= 2:
            ws.cell(row=1, column=5, value="bin_label").font = Font(bold=True)
            ws.cell(row=1, column=5).alignment = Alignment(horizontal="center")
            for row in range(2, ws.max_row + 1):
                bin_left = ws.cell(row=row, column=1).value
                bin_right = ws.cell(row=row, column=2).value
                if bin_left is None or bin_right is None:
                    label = None
                else:
                    label = f"[{float(bin_left):.3g}, {float(bin_right):.3g})"
                ws.cell(row=row, column=5, value=label)
            ws.column_dimensions["E"].width = 18

            chart = BarChart()
            chart.type = "col"
            chart.title = "Histogram"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Bin"
            data = Reference(ws, min_col=3, max_col=3, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=5, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 20
            chart.height = 12
            ws.add_chart(chart, "F2")

    wb.save(out_xlsx)
    return out_xlsx
